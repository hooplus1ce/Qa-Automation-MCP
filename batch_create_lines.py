import json
import time
import openpyxl
import httpx
from playwright.sync_api import sync_playwright

def get_auth_from_browser():
    """方式 A: 从 9222 端口 Chrome 实时提取 token 与 cookie"""
    print("正在连接 9222 端口浏览器提取实时鉴权凭证...")
    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp("http://127.0.0.1:9222")
        context = browser.contexts[0]
        page = [pg for pg in context.pages if "demo18-scm" in pg.url][0]
        
        # Extract cookies
        cookies = context.cookies(["https://demo18-scm.hoolinks.com"])
        cookie_header = "; ".join([f"{c['name']}={c['value']}" for c in cookies])
        
        # Extract token from storage / frame
        token = page.evaluate("""() => {
            return localStorage.getItem('token') || 
                   sessionStorage.getItem('token') ||
                   '0e5638ef-9349-41ac-9ecf-9472f75adfab1788574265760';
        }""")
        
        # Also check if token is in cookie
        for c in cookies:
            if c['name'] in ['HL-Access-Token', 'cookie_token', 'UCTOKEN']:
                token = c['value']
                break
                
        print(f"提取成功! Token: {token[:12]}... Cookie 长度: {len(cookie_header)}")
        return token, cookie_header

def batch_create():
    excel_file = "生和堂食品APS批量产线导入数据表.xlsx"
    print(f"正在加载产线数据文件: {excel_file}...")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["批量产线主数据表"]
    
    total_rows = ws.max_row - 1
    print(f"共发现 {total_rows} 条待创建产线记录。")
    
    # 提取凭证
    token, cookie_header = get_auth_from_browser()
    
    headers = {
        "accept": "*/*",
        "accept-language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7",
        "content-type": "application/json",
        "hl-access-token": token,
        "cookie": cookie_header,
        "origin": "https://demo18-scm.hoolinks.com",
        "referer": "https://demo18-scm.hoolinks.com/static/spo/productionLineManagement/form?mode=create"
    }
    
    create_url = "https://demo18-scm.hoolinks.com/scmpsm/aps/productionLine/create"
    
    success_records = []
    failed_records = []
    
    start_time = time.time()
    
    with httpx.Client(timeout=15.0) as client:
        for row_idx in range(2, ws.max_row + 1):
            line_idx = row_idx - 1
            # 找到 apiPayload 列（第 24 列）
            payload_str = ws.cell(row_idx, 24).value
            if not payload_str:
                # 兼容第 23 列
                payload_str = ws.cell(row_idx, 23).value
                
            payload = json.loads(payload_str)
            line_code = payload["lineCode"]
            line_name = payload["lineName"]
            
            try:
                resp = client.post(create_url, json=payload, headers=headers)
                res_data = resp.json()
                
                if res_data.get("ok") or res_data.get("status") == 0:
                    line_id = res_data.get("data")
                    success_records.append({"lineCode": line_code, "lineName": line_name, "id": line_id})
                    print(f"[{line_idx:03d}/{total_rows:03d}] [成功] {line_code} | {line_name} -> ID: {line_id}")
                else:
                    err_msg = res_data.get("msg") or str(res_data)
                    failed_records.append({"lineCode": line_code, "lineName": line_name, "error": err_msg})
                    print(f"[{line_idx:03d}/{total_rows:03d}] [失败] {line_code} | {line_name} -> 原因: {err_msg}")
            except Exception as e:
                failed_records.append({"lineCode": line_code, "lineName": line_name, "error": str(e)})
                print(f"[{line_idx:03d}/{total_rows:03d}] [异常] {line_code} | {line_name} -> 异常: {e}")
                
            # 控制请求速率，避免触发后端限流
            time.sleep(0.05)
            
    elapsed = time.time() - start_time
    
    print("\n" + "="*50)
    print(f"批量执行完成! 耗时: {elapsed:.2f} 秒")
    print(f"总计: {total_rows} 条 | 成功: {len(success_records)} 条 | 失败: {len(failed_records)} 条")
    print("="*50)
    
    if failed_records:
        print("\n失败记录明细:")
        for fr in failed_records:
            print(f"  {fr['lineCode']}: {fr['error']}")
            
    return len(success_records), len(failed_records)

if __name__ == "__main__":
    batch_create()
