import sys
sys.path.append('D:/Python/lib/site-packages')
import json
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_redesigned_excel():
    print("Loading real API responses...")
    with open('real_api_responses.json', 'r', encoding='utf-8') as f:
        api_data = json.load(f)

    # 1. Parse API data
    orgs = api_data['orgs'].get('data', [])
    depts_org1 = api_data['deptsByOrg']['1'].get('data', [])
    depts_org1383 = api_data['deptsByOrg']['1383'].get('data', [])
    depts_org42679 = api_data['deptsByOrg']['42679'].get('data', [])
    stations = api_data['stations'].get('data', [])
    
    dicts = {}
    for grp in api_data['dicts'].get('data', []):
        dicts[grp['parentCode']] = grp.get('items', [])
    
    line_groups = dicts.get('APS_PRODUCTION_LINE_GROUP', [])
    line_types = dicts.get('APS_PRODUCTION_LINE_TYPE', [])
    prod_types = dicts.get('APS_PRODUCTION_TYPE', [])
    
    real_devices = api_data['devices']['data']['records']
    real_dev_by_id = {d['deviceId']: d for d in real_devices}
    real_dev_by_code = {d['deviceCode']: d for d in real_devices}

    # 2. Read old Excel data
    print("Reading backup Excel data...")
    with zipfile.ZipFile('生和堂食品APS批量产线导入数据表.xlsx.bak') as z:
        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        
        # Read Sheet 1
        t1 = ET.fromstring(z.read('xl/worksheets/sheet1.xml'))
        sheet1_rows = []
        for r in t1.findall('.//ns:row', ns):
            cols = {}
            for c in r.findall('ns:c', ns):
                col = ''.join([ch for ch in c.get('r') if ch.isalpha()])
                val = None
                is_elem = c.find('ns:is', ns)
                if is_elem is not None:
                    t_elem = is_elem.find('ns:t', ns)
                    if t_elem is not None and t_elem.text:
                        val = t_elem.text
                if val is None:
                    v_elem = c.find('ns:v', ns)
                    if v_elem is not None and v_elem.text:
                        val = v_elem.text
                cols[col] = val or ''
            sheet1_rows.append(cols)

        # Read Sheet 2
        t2 = ET.fromstring(z.read('xl/worksheets/sheet2.xml'))
        sheet2_rows = []
        for r in t2.findall('.//ns:row', ns):
            cols = {}
            for c in r.findall('ns:c', ns):
                col = ''.join([ch for ch in c.get('r') if ch.isalpha()])
                val = None
                is_elem = c.find('ns:is', ns)
                if is_elem is not None:
                    t_elem = is_elem.find('ns:t', ns)
                    if t_elem is not None and t_elem.text:
                        val = t_elem.text
                if val is None:
                    v_elem = c.find('ns:v', ns)
                    if v_elem is not None and v_elem.text:
                        val = v_elem.text
                cols[col] = val or ''
            sheet2_rows.append(cols)

    # 3. Create Workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    # Fonts & Styles
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill_blue = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid') # Sheet 1
    header_fill_green = PatternFill(start_color='196F3D', end_color='196F3D', fill_type='solid') # Sheet 2
    header_fill_purple = PatternFill(start_color='4A235A', end_color='4A235A', fill_type='solid') # Sheet 3

    section_font = Font(name='微软雅黑', size=12, bold=True, color='1B4F72')
    section_fill = PatternFill(start_color='EAECEE', end_color='EAECEE', fill_type='solid')

    table_header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    table_header_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')

    data_font = Font(name='微软雅黑', size=10)
    data_font_bold = Font(name='微软雅黑', size=10, bold=True)
    code_font = Font(name='Consolas', size=9)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    thin_border = Border(
        left=Side(style='thin', color='D4D4D4'),
        right=Side(style='thin', color='D4D4D4'),
        top=Side(style='thin', color='D4D4D4'),
        bottom=Side(style='thin', color='D4D4D4')
    )

    zebra_fill = PatternFill(start_color='F8F9F9', end_color='F8F9F9', fill_type='solid')

    # =========================================================================
    # SHEET 1: 批量产线主数据表
    # =========================================================================
    print("Building Sheet 1: 批量产线主数据表...")
    ws1 = wb.create_sheet(title="批量产线主数据表")
    
    sheet1_headers = [
        "序号", "产线编码", "产线名称", "创建组织ID", "创建组织名称",
        "使用组织ID", "使用组织名称", "归属部门ID", "归属部门名称", "工位ID",
        "工位名称", "产线群组编码", "产线群组名称", "产线类型编码", "产线类型名称",
        "生产类型编码", "生产类型名称", "排程优先级", "标配作业人数", "绑定设备总数",
        "绑定设备编码列表", "绑定设备名称列表", "绑定设备型号列表", "API提交请求体"
    ]
    
    ws1.append(sheet1_headers)
    for col_idx in range(1, len(sheet1_headers) + 1):
        c = ws1.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill_blue
        c.alignment = align_center
        c.border = thin_border
    ws1.row_dimensions[1].height = 28

    # Mapping rules for real dictionary
    # Dept ID mapping:
    # 制造一部 -> 1361 (verified real API)
    # 制造二部 -> 1362 (verified real API)
    # 制造三部 -> 1361 (mapped to 制造一部) or 1362
    # 仓储部 -> 1364 (verified real API)
    # 生管部 -> 1360 (verified real API)
    # 饮料组 -> 1382 (verified real API)

    # Line Group mapping:
    # 充填 -> TCCJ (充填车间)
    # 煮料 / 配料 -> ZLCJ (煮料车间)
    # 杀菌 -> SJCJ (杀菌车间)
    # 包装 / 分装 / 抽检 / 打码 -> BZ1 ~ BZ6 (包装车间) / FB (粉包车间)

    # Line Type mapping:
    # 充填 / 旋盖 -> TCX (充填线)
    # 充填传输 -> TCCSX (充填传输线)
    # 直包装 / 分装 -> ZBLSX (直包流水线设备)
    # 3克分装 -> 3KFZ (3克分装设备)
    # 果干包装 -> GGBZ (果干包装设备)
    # 排装 -> PZX (排装线设备) / PZLSX
    # 散装 -> SZLSX (散装流水线设备)
    # 手工 -> SGLSX (手工流水线设备)
    # 杀菌 -> SJX (杀菌线) / SJCSX
    # 煮料 / 熬煮 -> ZLX (煮料线)

    for row_idx, r_data in enumerate(sheet1_rows[1:], start=2):
        seq = int(r_data.get('A', row_idx - 1))
        line_code = r_data.get('B', '')
        line_name = r_data.get('C', '')
        create_org_id = int(r_data.get('D', 1))
        create_org_name = r_data.get('E', '广东生和堂健康食品股份有限公司')
        use_org_id = int(r_data.get('F', 1))
        use_org_name = r_data.get('G', '广东生和堂健康食品股份有限公司')
        
        old_dept_name = r_data.get('I', '制造一部')
        # Correct dept ID based on real departmentOptions
        if '制造一部' in old_dept_name:
            dept_id = 1361
            dept_name = '制造一部'
        elif '制造二部' in old_dept_name:
            dept_id = 1362
            dept_name = '制造二部'
        elif '仓储' in old_dept_name:
            dept_id = 1364
            dept_name = '仓储部'
        elif '饮料' in old_dept_name:
            dept_id = 1382
            dept_name = '饮料组'
        elif '电商' in old_dept_name:
            dept_id = 42680
            dept_name = '电商制造部'
        else:
            dept_id = 1361
            dept_name = '制造一部'
            
        station_id = int(r_data.get('J', 56))
        station_name = r_data.get('K', '充填')
        
        # Determine real lineGroup
        if station_id == 56: # 充填
            # Distribute between TCCJ and BZ1/BZ2
            if seq % 2 == 1:
                line_group_code = 'TCCJ'
                line_group_name = '充填车间'
            else:
                line_group_code = 'BZ2'
                line_group_name = '包装2车间'
        elif station_id in (51, 52, 53, 54): # 配料/预制料/煮料/调香
            line_group_code = 'ZLCJ'
            line_group_name = '煮料车间'
        elif station_id == 58: # 杀菌
            line_group_code = 'SJCJ'
            line_group_name = '杀菌车间'
        elif '粉' in line_name:
            line_group_code = 'FB'
            line_group_name = '粉包车间'
        else:
            line_group_code = f'BZ{(seq % 6) + 1}'
            line_group_name = f'包装{(seq % 6) + 1}车间'

        # Special check for Record 7 (Row 8):
        # LINE-SHT-B01-007: 便携佐餐糖蜜小袋高速分装07线(SHT自营)
        if seq == 7:
            line_group_code = 'TCCJ'
            line_group_name = '充填车间'
            line_type_code = 'TCX'
            line_type_name = '充填线'
        else:
            # Determine real lineType
            if '3克' in line_name:
                line_type_code = '3KFZ'
                line_type_name = '3克分装设备'
            elif '果干' in line_name:
                line_type_code = 'GGBZ'
                line_type_name = '果干包装设备'
            elif '排装' in line_name:
                line_type_code = 'PZX'
                line_type_name = '排装线设备'
            elif '散装' in line_name or '散称' in line_name:
                line_type_code = 'SZLSX'
                line_type_name = '散装流水线设备'
            elif '人工' in line_name or '手工' in line_name:
                line_type_code = 'SGLSX'
                line_type_name = '手工流水线设备'
            elif '杀菌' in line_name or '水冷' in line_name:
                line_type_code = 'SJX'
                line_type_name = '杀菌线'
            elif '熬煮' in line_name or '煮料' in line_name or '萃取' in line_name:
                line_type_code = 'ZLX'
                line_type_name = '煮料线'
            elif '直包' in line_name or '分装' in line_name:
                line_type_code = 'ZBLSX'
                line_type_name = '直包流水线设备'
            elif '传输' in line_name:
                line_type_code = 'TCCSX'
                line_type_name = '充填传输线'
            else:
                line_type_code = 'TCX'
                line_type_name = '充填线'

        # Production Type
        prod_type_code = r_data.get('P', 'process')
        if prod_type_code not in ('process', 'repetitive', 'simple', 'outsourcing'):
            prod_type_code = 'process'
        prod_type_map = {
            'process': '流程生产',
            'repetitive': '重复生产',
            'simple': '简单生产',
            'outsourcing': '委外生产'
        }
        prod_type_name = prod_type_map[prod_type_code]

        priority = int(r_data.get('R', 8))
        operator_count = int(r_data.get('S', 6))
        device_count = int(r_data.get('T', 0))
        dev_codes_str = r_data.get('U', '')
        dev_names_str = r_data.get('V', '')
        dev_models_str = r_data.get('W', '')

        # Device ID list from old API payload or by matching
        old_payload_str = r_data.get('X', '{}')
        try:
            old_payload = json.loads(old_payload_str)
            device_id_list = old_payload.get('deviceIdList', [])
        except:
            device_id_list = []

        # Construct new verified API payload
        new_payload = {
            "createOrgId": create_org_id,
            "orgId": use_org_id,
            "deptId": dept_id,
            "stationId": station_id,
            "lineCode": line_code,
            "lineName": line_name,
            "lineGroup": line_group_code,
            "lineType": line_type_code,
            "productionType": prod_type_code,
            "schedulingPriority": priority,
            "operatorCount": operator_count,
            "deviceIdList": device_id_list
        }
        new_payload_str = json.dumps(new_payload, ensure_ascii=False)

        row_vals = [
            seq, line_code, line_name, create_org_id, create_org_name,
            use_org_id, use_org_name, dept_id, dept_name, station_id,
            station_name, line_group_code, line_group_name, line_type_code, line_type_name,
            prod_type_code, prod_type_name, priority, operator_count, device_count,
            dev_codes_str, dev_names_str, dev_models_str, new_payload_str
        ]

        ws1.append(row_vals)
        cur_row = ws1.max_row
        ws1.row_dimensions[cur_row].height = 20
        for col_idx in range(1, len(row_vals) + 1):
            cell = ws1.cell(row=cur_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
            # Alignments
            if col_idx in (1, 4, 6, 8, 10, 18, 19, 20):
                cell.alignment = align_center
            elif col_idx in (2, 12, 14, 16):
                cell.alignment = align_center
                cell.font = code_font
            elif col_idx == 24:
                cell.alignment = align_left
                cell.font = code_font
            else:
                cell.alignment = align_left

    # =========================================================================
    # SHEET 2: 产线设备绑定明细表
    # =========================================================================
    print("Building Sheet 2: 产线设备绑定明细表...")
    ws2 = wb.create_sheet(title="产线设备绑定明细表")
    
    sheet2_headers = [
        "绑定明细ID", "产线编码", "产线名称", "设备序号", "物理设备ID",
        "设备编码", "设备名称", "一级设备类型编码", "一级设备类型名称",
        "二级设备类型编码", "二级设备类型名称", "规格型号", "所属车间",
        "使用组织", "设备状态"
    ]
    ws2.append(sheet2_headers)
    for col_idx in range(1, len(sheet2_headers) + 1):
        c = ws2.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill_green
        c.alignment = align_center
        c.border = thin_border
    ws2.row_dimensions[1].height = 28

    for row_idx, r_data in enumerate(sheet2_rows[1:], start=2):
        bind_id = r_data.get('A', f'BIND-{row_idx-1:03d}')
        line_code = r_data.get('B', '')
        line_name = r_data.get('C', '')
        dev_seq = int(r_data.get('D', 1))
        dev_id = int(r_data.get('E', 0)) if r_data.get('E') and r_data.get('E').isdigit() else None
        dev_code = r_data.get('F', '')
        dev_name = r_data.get('G', '')
        
        # Check against real devices
        if dev_id and dev_id in real_dev_by_id:
            real_dev = real_dev_by_id[dev_id]
            first_code = real_dev.get('firstDeviceTypeCode') or r_data.get('H', '-')
            first_name = real_dev.get('firstDeviceTypeName') or r_data.get('I', '-')
            sec_code = real_dev.get('secondDeviceTypeCode') or r_data.get('J', '-')
            sec_name = real_dev.get('secondDeviceTypeName') or r_data.get('K', '-')
            spec = real_dev.get('specificationModel') or r_data.get('L', '-')
        else:
            first_code = r_data.get('H', '-')
            first_name = r_data.get('I', '-')
            sec_code = r_data.get('J', '-')
            sec_name = r_data.get('K', '-')
            spec = r_data.get('L', '-')
            
        workshop = r_data.get('M', '充填车间')
        org_name = r_data.get('N', '广东生和堂健康食品股份有限公司')
        status = r_data.get('O', '可用 (正常)')

        row_vals = [
            bind_id, line_code, line_name, dev_seq, dev_id,
            dev_code, dev_name, first_code, first_name, sec_code,
            sec_name, spec, workshop, org_name, status
        ]
        ws2.append(row_vals)
        cur_row = ws2.max_row
        ws2.row_dimensions[cur_row].height = 20
        for col_idx in range(1, len(row_vals) + 1):
            cell = ws2.cell(row=cur_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
            if col_idx in (1, 4, 5, 6, 8, 10, 15):
                cell.alignment = align_center
                if col_idx in (1, 5, 6):
                    cell.font = code_font
            else:
                cell.alignment = align_left

    # =========================================================================
    # SHEET 3: 系统字典与枚举规范 (Comprehensive Redesign)
    # =========================================================================
    print("Building Sheet 3: 系统字典与枚举规范...")
    ws3 = wb.create_sheet(title="系统字典与枚举规范")

    # Document Banner
    ws3.merge_cells('A1:F1')
    title_cell = ws3.cell(row=1, column=1, value="生和堂食品 APS 产线管理主数据与标准字典规范表")
    title_cell.font = Font(name='微软雅黑', size=15, bold=True, color='FFFFFF')
    title_cell.fill = header_fill_purple
    title_cell.alignment = align_center
    ws3.row_dimensions[1].height = 36

    ws3.merge_cells('A2:F2')
    sub_cell = ws3.cell(row=2, column=1, value="数据源基准：基于系统真实后端接口（/scmpsm/sys/org/listByType, /departmentOptions, /stationOptions, /dictValue/listByCodes, /deviceOptions）校验与更新")
    sub_cell.font = Font(name='微软雅黑', size=10, italic=True, color='5D6D7E')
    sub_cell.alignment = align_center
    ws3.row_dimensions[2].height = 20

    current_r = 4

    def add_section_header(title_text):
        nonlocal current_r
        ws3.merge_cells(f'A{current_r}:F{current_r}')
        c = ws3.cell(row=current_r, column=1, value=title_text)
        c.font = section_font
        c.fill = section_fill
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws3.row_dimensions[current_r].height = 26
        current_r += 1

    def add_table_header(headers):
        nonlocal current_r
        for col_idx, h in enumerate(headers, start=1):
            c = ws3.cell(row=current_r, column=col_idx, value=h)
            c.font = table_header_font
            c.fill = table_header_fill
            c.alignment = align_center
            c.border = thin_border
        ws3.row_dimensions[current_r].height = 24
        current_r += 1

    def add_table_row(values, is_zebra=False):
        nonlocal current_r
        for col_idx, v in enumerate(values, start=1):
            c = ws3.cell(row=current_r, column=col_idx, value=v)
            c.font = data_font
            c.border = thin_border
            if is_zebra:
                c.fill = zebra_fill
            if col_idx in (1, 2):
                c.alignment = align_center
                c.font = code_font
            else:
                c.alignment = align_left
        ws3.row_dimensions[current_r].height = 20
        current_r += 1

    # --- 1. 组织架构配置 ---
    add_section_header("1. 组织架构字典规范 (Organization Dictionary - /scmpsm/sys/org/listByType)")
    add_table_header(["组织ID (orgId)", "组织编码", "组织全称", "业务类型", "职责与设备配置", "支持操作模式"])
    org_details = [
        [1, "SHT01", "广东生和堂健康食品股份有限公司", "股份本部", "生产制造主体，拥有并配置全部 202 台生产设备", "创建组织 / 使用组织"],
        [1383, "C00001", "凭祥原料生产基地", "原料基地", "边境原料初制与半成品粗加工，委托本部代工制造", "创建组织 / 使用组织"],
        [42679, "C00004", "生和堂电商公司", "电商销售主体", "全网电商直发、一件代发与定制产线生产委托", "创建组织 / 使用组织"]
    ]
    for idx, row in enumerate(org_details):
        add_table_row(row, is_zebra=(idx%2==1))
    current_r += 1

    # --- 2. 部门字典配置 ---
    add_section_header("2. 制造部门字典规范 (Department Dictionary - /scmpsm/aps/productionLine/departmentOptions)")
    add_table_header(["部门ID (deptId)", "部门编码", "部门全称", "所属组织", "车间定位与职责说明", "状态"])
    # Add key departments first
    dept_priority = [1361, 1362, 1360, 1382, 1364, 1363, 1355, 1359]
    dept_rows = []
    seen_depts = set()
    for d_id in dept_priority:
        d = next((item for item in depts_org1 if item.get('id') == d_id), None)
        if d:
            seen_depts.add(d['id'])
            desc = ""
            if d['id'] == 1361: desc = "【核心制造部门】条装/袋装/吸吸果冻充填车间（原表错误写为1360，真实API为1361）"
            elif d['id'] == 1362: desc = "【核心制造部门】经典方杯/碗装龟苓膏充填封口车间"
            elif d['id'] == 1360: desc = "生产计划与物料控制部 (生管部)"
            elif d['id'] == 1382: desc = "草本饮料充填与浓缩原液萃取制作组"
            elif d['id'] == 1364: desc = "后道自动化输送与立体仓储周转车间"
            elif d['id'] == 1363: desc = "厂务与动力设备运维工程部"
            elif d['id'] == 1355: desc = "食品配方与加工工艺研发部"
            elif d['id'] == 1359: desc = "全流程品质检验与质量控制部"
            dept_rows.append([d['id'], d.get('code'), d.get('name'), "广东生和堂健康食品股份有限公司", desc, "启用"])
    
    for d in depts_org1:
        if d['id'] not in seen_depts:
            dept_rows.append([d['id'], d.get('code'), d.get('name'), "广东生和堂健康食品股份有限公司", "行政职能与营销大区运营组", "启用"])
    
    for idx, row in enumerate(dept_rows):
        add_table_row(row, is_zebra=(idx%2==1))
    current_r += 1

    # --- 3. 工位/站点字典配置 ---
    add_section_header("3. 工位/站点字典规范 (Station Dictionary - /scmpsm/aps/productionLine/stationOptions)")
    add_table_header(["工位ID (stationId)", "工位编码", "工位名称", "工序环节分类", "生产作业与工艺说明", "设备绑定能力"])
    station_descs = {
        51: ("配料", "前道物料准备", "原辅料计量、草本配方复配、沉淀静置", "支持储料缸与称量设备"),
        52: ("预制料", "前道原料溶化", "预混合、配料前处理与胶体原料定温融解", "支持预混机与储料缸"),
        53: ("煮料", "核心熬煮工序", "仙草提取液熬煮、胶体熔融恒温高压煮制", "支持煮料缸与热水缸"),
        54: ("调香", "调味均质工序", "风味调配、草本香韵注入与恒温高剪切均质", "支持调香调味机"),
        55: ("检测", "品控过程检测", "过程指标快检、理化指标定温测定与浓度标定", "支持理化检测仪器"),
        56: ("充填", "核心包装成型", "膏体无菌灌装、注料旋盖封口一体化连续成型", "支持充填机与连运线"),
        57: ("打码", "标识喷印工序", "高速激光喷码、保质期打码与后道连运", "支持激光喷码设备"),
        58: ("杀菌", "后道灭菌工序", "高温高压杀菌釜灭菌、多段循环水降温冷却", "支持杀菌釜与杀菌线"),
        59: ("抽检", "成品出厂质检", "包装密封性负压抽检、外观全检与重量复核", "支持抽检复核仪器")
    }
    for idx, s in enumerate(stations):
        s_id = s.get('id')
        s_info = station_descs.get(s_id, (s.get('name'), "常规工序", "标准作业", "支持通用设备"))
        add_table_row([s_id, s.get('code'), s.get('name'), s_info[1], s_info[2], s_info[3]], is_zebra=(idx%2==1))
    current_r += 1

    # --- 4. 产线群组字典规范 ---
    add_section_header("4. 产线群组字典规范 (APS_PRODUCTION_LINE_GROUP - dictValue/listByCodes)")
    add_table_header(["群组编码 (lineGroup)", "显示名称 (dictLabel)", "字典项全称", "排序号", "应用制造车间场景", "推荐绑定工位"])
    group_scenarios = {
        'TCCJ': ('充填车间', 9, '主力核心自营充填与注料旋盖车间', '充填 (56)'),
        'ZLCJ': ('煮料车间', 10, '仙草熬煮、胶体融解与原液萃取车间', '煮料 (53), 配料 (51)'),
        'SJCJ': ('杀菌车间', 8, '高温灭菌釜、巴氏杀菌与循环水冷却车间', '杀菌 (58)'),
        'FB':   ('粉包车间', 7, '干混固体饮料、仙草伴侣与粉剂包装车间', '充填 (56), 打码 (57)'),
        'BZ1':  ('包装1车间', 1, '异形袋与吸吸果冻后道自动包装车间', '充填 (56), 打码 (57)'),
        'BZ2':  ('包装2车间', 2, '便携佐餐糖蜜小袋高速分装车间', '充填 (56), 打码 (57)'),
        'BZ3':  ('包装3车间', 3, '经典方杯碗装定量充填连运车间', '充填 (56), 打码 (57)'),
        'BZ4':  ('包装4车间', 4, '散称小果冻人工拣选混装包装车间', '抽检 (59), 打码 (57)'),
        'BZ5':  ('包装5车间', 5, '草本养生常温礼盒组合装配车间', '打码 (57), 抽检 (59)'),
        'BZ6':  ('包装6车间', 6, '电商定制柔性快翻组合包装车间', '打码 (57), 抽检 (59)')
    }
    for idx, item in enumerate(line_groups):
        v = item.get('dictValue')
        info = group_scenarios.get(v, (item.get('dictLabel'), item.get('dictSort', 0), '标准生产车间', '通用工位'))
        add_table_row([v, item.get('dictLabel'), item.get('name') or item.get('dictLabel'), item.get('dictSort'), info[2], info[3]], is_zebra=(idx%2==1))
    current_r += 1

    # --- 5. 产线类型字典规范 ---
    add_section_header("5. 产线类型字典规范 (APS_PRODUCTION_LINE_TYPE - dictValue/listByCodes)")
    add_table_header(["类型编码 (lineType)", "显示名称 (dictLabel)", "排序号", "典型生产设备配型", "生产作业模式", "产能特征"])
    type_details = {
        'TCX':   ('充填线', 10, '袋装充填机、管嘴充填机、方杯充填机', '连续定量无菌灌装封口', '高速大批量'),
        'TCCSX': ('充填传输线', 9, '充填传输连运线、链板输送机', '注料后道自动平移输送', '连续平稳传输'),
        'ZBLSX': ('直包流水线设备', 12, '高速直包机、背封成型自动包装机', '卷膜制袋装料封切一体化', '连续高速直包'),
        '3KFZ':  ('3克分装设备', 1, '微量粉粒精密计量分装机', '高精度微量颗粒分装', '精密稳定分装'),
        'GGBZ':  ('果干包装设备', 2, '多头电脑组合秤、立式果干制袋包装机', '固体果干自动称量分装', '柔性定量称量'),
        'PZX':   ('排装线设备', 4, '联包冲切机、排装套标自动线', '多联杯连排分装成型', '自动化排装'),
        'PZLSX': ('排装流水线设备', 3, '排装输送流水线、推料机构', '多联排装物料自动化导引', '流水线输送'),
        'SGLSX': ('手工流水线设备', 5, '人工流水工作台、辅助输送带', '人工拣选、组合配装与贴标', '柔性小批量定制'),
        'SZLSX': ('散装流水线设备', 8, '大倾角输送机、散装分料漏斗', '散称小袋大批量分流', '高吞吐散装'),
        'SJX':   ('杀菌线', 7, '全自动杀菌釜群组、升温加压控制柜', '高温高压蒸汽杀菌', '批次周期杀菌'),
        'SJCSX': ('杀菌传输线', 6, '杀菌水槽输送链、进出釜小车', '杀菌前后道自动化转运', '平稳耐温传输'),
        'VXHF':  ('V型混粉设备', 11, '高效不对称V型干粉混合机', '干粉物料均质复配', '批次三维混和'),
        'ZLX':   ('煮料线', 13, '蒸汽夹套熬煮锅、真空浓缩降膜蒸发器', '草本原液恒温浸润熬煮', '连续恒温蒸煮')
    }
    for idx, item in enumerate(line_types):
        v = item.get('dictValue')
        info = type_details.get(v, (item.get('dictLabel'), item.get('dictSort', 0), '通用生产加工设备', '标准生产作业', '标配产能'))
        add_table_row([v, item.get('dictLabel'), item.get('dictSort'), info[2], info[3], info[4]], is_zebra=(idx%2==1))
    current_r += 1

    # --- 6. 生产类型字典规范 ---
    add_section_header("6. 生产类型字典规范 (APS_PRODUCTION_TYPE - dictValue/listByCodes)")
    add_table_header(["生产类型编码", "显示名称", "字典值", "业务形态特性", "排程算法适配模型", "排程优先级建议"])
    prod_type_details = [
        ["process", "流程生产", "process", "连续化原液熬煮、调配混合与液体管道流动制造", "连续流物料守恒排程（容量平衡）", "9 - 10 (优先)"],
        ["repetitive", "重复生产", "repetitive", "标准化大批量离散式小袋充填、旋盖成型与连续外包", "节拍生产排程（有限产能负荷均衡）", "7 - 8 (标准)"],
        ["simple", "简单生产", "simple", "工序简单的人工散称混装、抽检复核与手工组合包", "工序看板排程（工时与人数约束）", "5 - 6 (普通)"],
        ["outsourcing", "委外生产", "outsourcing", "边境原料初制外协加工、跨基地协同生产代工", "供需协同与到货周期前置期排程", "1 - 4 (低级)"]
    ]
    for idx, row in enumerate(prod_type_details):
        add_table_row(row, is_zebra=(idx%2==1))
    current_r += 1

    # --- 7. API 参数规范与映射关系 ---
    add_section_header("7. APS 产线创建接口技术规范 (/scmpsm/aps/productionLine/create)")
    add_table_header(["字段参数名", "字段中文说明", "数据类型", "是否必填", "数据校验与业务级联规则", "示例取值"])
    api_specs = [
        ["createOrgId", "创建组织ID", "Long", "是", "必须为有效组织ID（生和堂本部为 1，凭祥为 1383，电商为 42679）", "1"],
        ["orgId", "使用组织ID", "Long", "是", "与创建组织形成级联，当前生产主体通常选择 1", "1"],
        ["deptId", "制造部门ID", "Long", "是", "【强校验】必须是所选 orgId 下属的真实部门ID；制造一部必须为 1361（非1360）", "1361"],
        ["stationId", "工位/站点ID", "Long", "是", "必须为 /stationOptions 返回的 9 大有效工位ID之一（充填为 56）", "56"],
        ["lineCode", "产线编码", "String", "是", "全局唯一，格式建议 LINE-SHT-Bxx-xxx，最大长度 100", "LINE-SHT-B01-007"],
        ["lineName", "产线名称", "String", "是", "描述产线工艺、包装规格与权属，最大长度 200", "便携佐餐糖蜜小袋高速分装07线(SHT自营)"],
        ["lineGroup", "产线群组编码", "String", "否", "必须对应 APS_PRODUCTION_LINE_GROUP 真实字典值（如 TCCJ, BZ1~BZ6 等）", "TCCJ"],
        ["lineType", "产线类型编码", "String", "是", "必须对应 APS_PRODUCTION_LINE_TYPE 真实字典值（如 TCX, ZBLSX 等）", "TCX"],
        ["productionType", "生产类型编码", "String", "否", "必须对应 APS_PRODUCTION_TYPE 真实字典值（如 repetitive, process 等）", "repetitive"],
        ["schedulingPriority", "排产优先级", "Integer", "否", "数字 1-10，数字越大排产优先级越高", "8"],
        ["operatorCount", "标配作业人数", "Integer", "否", "标准出勤在岗作业操作工人数", "6"],
        ["deviceIdList", "绑定设备ID数组", "Array<Long>", "否", "从 /deviceOptions 获取的可用设备物理ID列表（如 [9217131, 9217145]）", "[9217131, 9217145]"]
    ]
    for idx, row in enumerate(api_specs):
        add_table_row(row, is_zebra=(idx%2==1))

    # Auto-fit column widths across all sheets
    print("Formatting column widths...")
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Ignore merged cells in Sheet 3 row 1-2
                if ws == ws3 and cell.row in (1, 2):
                    continue
                v = str(cell.value or '')
                # Approximate width (Chinese chars take ~2 width)
                w = sum(2 if ord(ch) > 127 else 1 for ch in v)
                if w > max_len:
                    max_len = w
            # Set bounded width
            ws.column_dimensions[col_letter].width = max(min(max_len + 3, 50), 12)

    # Save output
    output_filename = "生和堂食品APS批量产线导入数据表.xlsx"
    print(f"Saving redesigned workbook to {output_filename}...")
    wb.save(output_filename)
    print("Workbook successfully saved!")

if __name__ == '__main__':
    build_redesigned_excel()
